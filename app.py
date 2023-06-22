from flask import Flask, flash, request, redirect, url_for, render_template, send_file
from dotenv import load_dotenv
import openpyxl
from pathlib import Path
import networkx as nx
load_dotenv()
from os import getenv
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "09saf0g80as7d0agd07sh07f0h97dj"
input_file = Path(getenv("EXCEL_PATH")) if getenv("EXCEL_PATH") else Path("./org_chart.xlsx")
template_file = Path("./template.xlsx")

# see https://flask.palletsprojects.com/en/2.3.x/patterns/fileuploads/
UPLOAD_FOLDER = '/path/to/the/uploads'
ALLOWED_EXTENSIONS = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def validate(file):
    errors = []
    xl = openpyxl.load_workbook(file)

    ##check the workbook contains the right sheets
    expected_sheet_names = {"people", "teams", "team_roles"}
    if expected_sheet_names - set(xl.sheetnames):
        errors.append(f"sheet names must include {expected_sheet_names}. found {xl.sheetnames}")
        return errors

    # check the headings on each sheet
    expected_headers = dict()
    expected_headers['teams'] = ['short_name',	'full_name',	'parent',	'contact_person']
    expected_headers['people'] = ['name',	'job_title',	'home_team']
    expected_headers['team_roles'] = ['job_title',	'job_type']

    for sheet in expected_sheet_names:
        sheet_headers = [cell.value for cell in xl[sheet][1]]
        if sheet_headers != expected_headers[sheet]:
            errors.append(f"Error - column headers on `{sheet}` sheet must be {expected_headers[sheet]}. Found {sheet_headers} instead")
    if errors:
        return errors

    ##check there are no cycles in the org chart tree
    G = nx.Graph()
    worksheet = xl['teams']
    # first add all the nodes to the graph
    for row_number in range(2, worksheet.max_row + 1):
        cell_name = f"A{row_number}"
        team_name = worksheet[cell_name].value
        if team_name:
            G.add_node(team_name)

    # first add all the nodes to the graph
    for row_number in range(2, worksheet.max_row + 1):
        team_cell_name = f"A{row_number}"
        parent_cell_name = f"C{row_number}"
        team_name = worksheet[team_cell_name].value
        parent_name = worksheet[parent_cell_name].value
        if team_name and parent_name:
            G.add_edge(team_name, parent_name)

    try:
        cycle = nx.find_cycle(G)
        if cycle:
            errors.append(f"Error - found circular references in Teams parent ({cycle})")
    except Exception as e:
        print(e)

    return errors


@app.route('/upload_file', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        try:
            errors = validate(file)
        except Exception as e:
            return ["error validating file"]

        if errors:
            return errors
        if file and allowed_file(file.filename):
            file.save(input_file)
            return redirect(url_for('index', name=file))
    return render_template('upload.html')


@app.route('/download_file', methods=['GET'])
def download_file():
    return send_file(input_file)


@app.route('/download_template', methods=['GET'])
def download_template():
    return send_file(input_file)


@app.route("/")
def index():
    return render_template('index.html')
